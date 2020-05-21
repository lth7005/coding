# -*- coding: utf-8 -*-
"""

파이썬으로 엑셀 다루기1 : 엑셀 파일 불러오기 및 데이터 출력

Created on Tue May 19 20:09:57 2020

@author: lth7000

@참고자료: (블로그)https://myjamong.tistory.com/51?category=853676

"""

#openpyxl모듈을 가져온다. 엑셀을 다룰 수 있게 해준다.
#가져오려면 명령프롬프트에서 pip install openpyxl 명령어로 설치해야 함.  
import openpyxl as opx

#(test1)엑셀 파일을 불러온다.
wb = opx.load_workbook('test.xlsx', data_only=True) 
#data_only=True를 하면 수식이 계산되어 값으로 반환. 안쓰면 수식으로 반환.
'''
>>> type(wb)
<class 'openpyxl.workbook.workbook.Workbook'>
'''
print('(test1) wb\n', wb, '\n') #이하 print('(test)~') 는 확인용 print

#(test2)엑셀 파일의 sheet를 불러온다.
ws = wb['성적산출']
'''
>>> type(ws)
<class 'openpyxl.worksheet.worksheet.Worksheet'>
'''
print('(test2) ws\n', ws, '\n')

#(test3-1)셀 출력
print("(test3-1) ws['A1']\n", ws["A1"], '\n')

#(test3-2) 셀 주소로 데이터 출력
print("(test3-2) ws['A1'].value\n", ws['A1'].value, '\n') #1 row, 1 column

#(test4) 셀 주소로 데이터 출력
print("(test4) ws['F3'].value\n", ws['F3'].value, '\n') #3 row, 6 column

#(test5) 셀 좌표로 데이터 출력
print('(test5) ws.cell(3,6).value\n', ws.cell(3,6).value, '\n') #3 row, 6 column

#.value를 붙여주면 cell 안의 값을 반환하고, 안쓰면 그 값이 저장된 상자?객체?주소?를 반환하는 것 같다.

#(test6) 시트에서 슬라이싱으로 작업할 부분 떼어 내기. A4부터 F6까지
ws_part = ws['A4':'F6'] #시트(ws)를 슬라이싱으로 떼어 내면(ws_part) .rows는 사용 못하지만 row 구분은 여전히 된다.-> 61번줄에서 a변수에 row단위로 저장된다.
                        #시트를 슬라이싱으로 떼어 내면 튜플로 저장된다.
                        #.rows는 시트 전체의 row를 불러올 때 사용하는 것 같다. 
                        #시트 전체에 빈 값(none)이 많아서 여기서는 시트 일부만 떼어서 사용하기로.
'''
>>> type(ws_part)
<class 'tuple'>
'''
print("(test6) ws['A4':'F6']\n", ws_part) #확인용 print

i=4; j=1; k=1 
for a in ws_part: #a는 ws_part의 각 row
    print('\n(test6-{}) {}열의 원소들\n{} \n'.format(k,i,a))
    k+=1
    for b in a: #b는 a(ws_part의 각 row)의 각 cell
        #print('(test6-%d)' %k, i, '열', j, '행의 원소', b, '는', b.value)
        print('(test6-{}) {}열 {}행의 원소 {}는 {}'.format(k, i, j, b, b.value))
        j+=1; k+=1
    i+=1; j=1

#(test7) 떼어 낸 부분의 데이터를 리스트로 저장하고 출력하기
print('\n(test7) 떼어 낸 부분의 데이터를 리스트로 저장하고 출력하기')
data = []
for row in ws_part: #ws_part는 슬라이싱으로 특정 부분을 떼어 낸 것. '.rows'는 시트 원본에만 사용하는 것 같다.
                    #row에는 시트에서 떼어 낸 부분인 ws_part의 row가 차례로 저장. 이것도 편한 것 같다.
    data_temp = []
    for cell in row:
        data_temp.append(cell.value)
    data.append(data_temp)
print(data)

'''
#none값이 너무 많다.
#none 출력 생략하는 방법?
c=[]; d=[];
for a in ws.rows: #ws는 '성적산출' 시트, ws.rows는 시트의 열들
    for b in a:
        c.append(b.value)
    d.append(c)
print(d)
'''
